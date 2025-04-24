import os
import openpyxl
from kivy.app import App 
from kivy.uix.screenmanager import Screen, ScreenManager
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from kivy.uix.button import Button
from kivy.uix.spinner import Spinner
from kivy.uix.scrollview import ScrollView
from kivy.core.window import Window
from kivy.uix.image import Image
from kivy.graphics import Color, Rectangle
from datetime import datetime
import sys

# Enhanced resource path function
def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    path = os.path.join(base_path, *relative_path.split('/'))
    path = os.path.normpath(path)
    
    if not os.path.exists(path):
        raise FileNotFoundError(f"Resource not found: {path}")
    
    return path

# Atur ukuran window
Window.size = (800, 600)

# Warna UI
BUTTON_COLOR = (0.184, 0.322, 0.627, 1)  # Warna navbar (#2f52a0)
RED_COLOR = (1, 0, 0, 1)  # Warna merah untuk peringatan
BLACK_COLOR = (0, 0, 0, 1)  # Warna hitam
WHITE_COLOR = (1, 1, 1, 1)  # Warna putih

class Navbar(BoxLayout):
    def __init__(self, back_callback=None, **kwargs):
        super().__init__(**kwargs)
        self.orientation = 'horizontal'
        self.size_hint_y = 0.1
        self.padding = [20, 10]
        self.spacing = 15

        with self.canvas.before:
            Color(*BUTTON_COLOR)
            self.rect = Rectangle(pos=self.pos, size=self.size)
        self.bind(pos=self.update_rect, size=self.update_rect)

        back_btn = Button(size_hint=(None, None), size=(65, 65), pos_hint={'center_y': 0.5},
                          background_normal='assets/icons/back_icon.png', 
                          background_down='assets/icons/back_icon.png')
        if back_callback:
            back_btn.bind(on_release=back_callback)
        self.add_widget(back_btn)

        logo = Image(source='assets/logo_unair.png', size_hint=(None, None), 
                     size=(50, 50), pos_hint={'center_y': 0.5})
        self.add_widget(logo)

        title_label = Label(text="RUANG BACA FAKULTAS SAINS DAN TEKNOLOGI", 
                          font_size=20, color=WHITE_COLOR, 
                          halign="left", valign="middle")
        title_label.bind(size=title_label.setter('text_size'))
        self.add_widget(title_label)

    def update_rect(self, *args):
        self.rect.pos = self.pos
        self.rect.size = self.size

class InputScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)

        with self.canvas.before:
            Color(*WHITE_COLOR)  # Latar belakang putih
            self.rect = Rectangle(pos=self.pos, size=self.size)
        self.bind(pos=self.update_rect, size=self.update_rect)

        self.selected_file = "data_mahasiswa.xlsx"
        self.warning_label = Label(text="", font_size=14, color=RED_COLOR, 
                                 size_hint=(1, 0.05), halign="center")

        main_layout = BoxLayout(orientation="vertical", padding=[20, 10], spacing=10)
        
        # Navbar
        navbar = Navbar(back_callback=self.go_back)
        main_layout.add_widget(navbar)

        # Judul
        title_layout = BoxLayout(size_hint=(1, 0.12), padding=[0, 10])
        title = Label(text="INPUT DATA MAHASISWA", font_size=24, 
                     size_hint=(1, 1), color=BLACK_COLOR, 
                     halign="center", valign="middle")
        title.bind(size=lambda instance, value: setattr(instance, 'text_size', value))
        title_layout.add_widget(title)
        main_layout.add_widget(title_layout)

        # Form Input
        scroll = ScrollView(size_hint=(1, 0.7))
        form_layout = GridLayout(cols=2, spacing=15, padding=[20, 10], 
                               size_hint_y=None, row_default_height=50)
        form_layout.bind(minimum_height=form_layout.setter('height'))

        labels = ["TANGGAL PENCATATAN", "NO. INDUK", "PENGARANG", "NIM", 
                 "TAHUN TERBIT", "JUDUL"]
        self.inputs = {}
        
        # Spinner Kategori
        form_layout.add_widget(Label(text="Jenis Karya Ilmiah", color=BLACK_COLOR))
        self.spinner_kategori = Spinner(text="Pilih", values=list(PROGRAM_STUDI.keys()),
                                      size_hint_y=None, height=40)
        self.spinner_kategori.bind(text=self.update_jenjang)
        form_layout.add_widget(self.spinner_kategori)
        
        # Spinner Jenjang
        form_layout.add_widget(Label(text="Jenjang", color=BLACK_COLOR))
        self.spinner_jenjang = Spinner(text="Pilih", values=[], 
                                      size_hint_y=None, height=40)
        self.spinner_jenjang.bind(text=self.update_prodi)
        form_layout.add_widget(self.spinner_jenjang)
        
        # Spinner Program Studi
        form_layout.add_widget(Label(text="Program Studi", color=BLACK_COLOR))
        self.spinner_prodi = Spinner(text="Pilih", values=[], 
                                    size_hint_y=None, height=40)
        form_layout.add_widget(self.spinner_prodi)
        
        # Input fields
        for label_text in labels:
            form_layout.add_widget(Label(text=label_text, color=BLACK_COLOR))
            input_field = TextInput(multiline=False, halign="center", 
                                  foreground_color=BLACK_COLOR,
                                  size_hint_y=None, height=40)
            input_field.bind(on_text_validate=self.focus_next_input)
            form_layout.add_widget(input_field)
            self.inputs[label_text] = input_field

        scroll.add_widget(form_layout)
        main_layout.add_widget(scroll)
        main_layout.add_widget(self.warning_label)

        # Tombol
        button_layout = BoxLayout(size_hint=(1, 0.1), spacing=20, padding=[20, 10])
        clear_btn = Button(text="CLEAR", background_color=BUTTON_COLOR, 
                          color=WHITE_COLOR, size_hint_y=None, height=40)
        submit_btn = Button(text="SUBMIT", background_color=BUTTON_COLOR, 
                           color=WHITE_COLOR, size_hint_y=None, height=40)
        clear_btn.bind(on_release=self.clear_form)
        submit_btn.bind(on_release=self.submit_form)
        button_layout.add_widget(clear_btn)
        button_layout.add_widget(submit_btn)
        main_layout.add_widget(button_layout)

        self.add_widget(main_layout)

    def update_rect(self, *args):
        self.rect.pos = self.pos
        self.rect.size = self.size

    def update_jenjang(self, spinner, text):
        self.spinner_jenjang.values = list(PROGRAM_STUDI.get(text, {}).keys())
        self.spinner_jenjang.text = "Pilih"
        self.spinner_prodi.text = "Pilih"
        self.spinner_prodi.values = []

    def update_prodi(self, spinner, text):
        kategori = self.spinner_kategori.text
        self.spinner_prodi.values = PROGRAM_STUDI.get(kategori, {}).get(text, [])
        self.spinner_prodi.text = "Pilih"

    def clear_form(self, instance):
        for input_field in self.inputs.values():
            input_field.text = ""
        self.spinner_kategori.text = "Pilih"
        self.spinner_jenjang.text = "Pilih"
        self.spinner_prodi.text = "Pilih"
        self.warning_label.text = "Form dibersihkan."

    def submit_form(self, instance):
        missing_fields = [label for label, input_field in self.inputs.items() 
                        if not input_field.text.strip()]
        
        if self.spinner_kategori.text == "Pilih" or \
           self.spinner_jenjang.text == "Pilih" or \
           self.spinner_prodi.text == "Pilih":
            missing_fields.append("Kategori, Jenjang, atau Program Studi")
        
        if missing_fields:
            self.warning_label.text = f"Silakan lengkapi terlebih dahulu: {', '.join(missing_fields)}"
        else:
            self.save_to_excel()
            self.warning_label.text = "Data berhasil disimpan."

    def save_to_excel(self):
        file_path = self.selected_file
        sheet_name = f"{self.spinner_kategori.text.upper()} {self.spinner_jenjang.text} {self.spinner_prodi.text}"
        
        if not os.path.exists(file_path):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = sheet_name
            sheet.append([sheet_name])
            sheet.append([])
            sheet.append(list(self.inputs.keys()))
            workbook.save(file_path)
        else:
            workbook = openpyxl.load_workbook(file_path)
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.create_sheet(sheet_name)
                sheet.append([sheet_name])
                sheet.append([])
                sheet.append(list(self.inputs.keys()))
        
        row_data = [self.inputs[label].text for label in self.inputs]
        sheet.append(row_data)

        # Menyesuaikan lebar kolom
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column].width = adjusted_width

        workbook.save(file_path)

    def go_back(self, instance):
        self.manager.current = 'home'

    def focus_next_input(self, instance):
        input_fields = list(self.inputs.values())
        current_index = input_fields.index(instance)
        
        if current_index < len(input_fields) - 1:
            input_fields[current_index + 1].focus = True
        else:
            self.spinner_kategori.focus = True

# Data Program Studi
PROGRAM_STUDI = {
    "Skripsi": {
        "S1": ["KIMIA", "BIOLOGI", "MATEMATIKA", "FISIKA", "TEKNIK BIOMEDIS", 
               "SISTEM INFORMASI", "STATISTIKA", "TEKNIK LINGKUNGAN"],
        "S2": ["KIMIA", "BIOLOGI", "MATEMATIKA", "TEKNIK BIOMEDIS"],
        "S3": ["KIMIA", "BIOLOGI", "MATEMATIKA", "FISIKA"]
    },
    "Tesis": {
        "S1": ["KIMIA", "BIOLOGI", "MATEMATIKA", "FISIKA", "TEKNIK BIOMEDIS", 
               "SISTEM INFORMASI", "STATISTIKA", "TEKNIK LINGKUNGAN"],
        "S2": ["KIMIA", "BIOLOGI", "MATEMATIKA", "TEKNIK BIOMEDIS"],
        "S3": ["KIMIA", "BIOLOGI", "MATEMATIKA", "FISIKA"]
    },
    "Disertasi": {
        "S1": ["KIMIA", "BIOLOGI", "MATEMATIKA", "FISIKA", "TEKNIK BIOMEDIS", 
               "SISTEM INFORMASI", "STATISTIKA", "TEKNIK LINGKUNGAN"],
        "S2": ["KIMIA", "BIOLOGI", "MATEMATIKA", "TEKNIK BIOMEDIS"],
        "S3": ["KIMIA", "BIOLOGI", "MATEMATIKA", "FISIKA"]
    }
}

class RuangBacaApp(App):
    def build(self):
        sm = ScreenManager()
        sm.add_widget(InputScreen(name="input"))  
        return sm

if __name__ == '__main__':
    RuangBacaApp().run()